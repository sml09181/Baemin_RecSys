# 참고: https://ssseoyneee.tistory.com/4

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import pandas as pd
from openpyxl import load_workbook, Workbook
import time
import os
import re
import json
import glob

# 음식점 아닌 카테고리 제거
etc = ['오락시설', '한약국,한약방', '국토교통부', '영화관', '화장품,향수', '장소대여', 'PC방',
       '문구,팬시용품', '네일아트,네일샵', '관리,안내', '침구,커튼', '출력,복사,제본',
       '마사지, 지압', '체험여행', '특산물,관광민예품', '남성정장', '찜질방', '운세,사주', 
       '고속도로휴게소', '웨딩컨설팅,플래너', '슈퍼,마트', '종합생활용품', 
       '기업', '공영주차장', '다이어트,비만', '주차장', '공방', '피부과', 
       '약국', '귀금속,시계', '가구', '폐업했거나 정보 제공이 중지된 장소', 
       '속눈썹증모,연장', '박물관', '성형외과', '스포츠용품', '화물운송', 
       'S', '안경원', '치과', '없음', '건강기능보조식품', '결혼예물', '애견용품', 
       '셀프,대여스튜디오', '동물병원', '드럭스토어', '왁싱,제모', '호텔', '종합패션', 
       '미용실', '공사,공단', '헬스장', '내과', '편의점', '미술관', '슈퍼,마트', 
       'GS칼텍스주유소', '전기,가스,수도사업', 'HD현대오일뱅크주유소', '약국', 
       '영화관', '종합가전', 'SK주유소', '고속도로휴게소', '종합생활용품', 
       '소프트웨어개발', '관람,체험', '유아,아동용품', '솔루션개발', '정장', 
       '남성의류', '신발', '유아동복', '상가,아케이드', '편의점', '화장실', 
       'S', '내과', '가방,핸드백', '등산,아웃도어', '카드단말기', '렌터카', '패션',
       '절,사찰', '은행', '캐주얼웨어', '스포츠용품', '주차장', '판촉,기념품', 
       '문화센터', '장례식장', '문구,팬시용품', '호텔', '안경원', '여성의류', 
       '드럭스토어', '서점', '주유소', '민간자동차검사소', '산업용품', '치과', 
       '현금인출기', '갤러리,화랑', '도장', '만화방', '공영주차장', '전문건설업', 
       '국토교통부', '특산물,관광민예품', '타이어,휠', '국제,항공화물', '휴대폰수리', 
       '기업', '관리,안내', '정형외과', '도보코스', '콘도,리조트', '해수욕장,해변', 
       '온천,스파', '기부,모금단체', '자연,생태공원', '세탁', '우편취급소', '차', 
       '노래방', '세탁소', '전망대', '회관', '한의원', '패션잡화', '의류제작', 
       '조명,디스플레이', '명절 무료 개방 주차장', '신용조합', '산부인과', 
       '병원,의원', '콘택트렌즈전문', '정육점', '교습학원,교습소', '컨벤션센터', 
       '당구장']

options = Options()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

user_URL_list = []
json_files = ['C:\\Recsys\\reviews\\1207_review.json'] # 음식점 크롤링 data가 들어있는 json 파일
file_name = 'Reviewer.xlsx'
sheet_number = os.path.basename(json_files[0]).split("_")[0]

try:
    # 기존 파일 열기
    xlsx = load_workbook(file_name)
except FileNotFoundError:
    # 파일 없으면 새로 생성
    xlsx = Workbook()

list_sheet = xlsx.create_sheet(f'output{sheet_number}_')
list_sheet.append(['user_id', 'nickname', 'store_id', 'store_name', 'store_location', 'category', 'content', 'date', 'num_visit', 'visit_time', 'url'])
xlsx.save(file_name)

user_id = -1
print(json_files)

for json_file_path in json_files:
    with open(json_file_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)

    df = pd.DataFrame(json_data)
    df["i_tags"] = df["i_tags"].apply(lambda x: ", ".join(x))     

    # 파일명에서 숫자만 추출
    file_number = os.path.basename(json_file_path).split("_")[0]
    
    print("data", json_file_path, file_number)

    start_index = 0
    # 아래 세 줄은 중간에 멈췄을 때를 대비한 코드
    # start_name = "skd****"
    # start_index = df[df.iloc[:, 0] == start_name].index[0] + 1
    # print(start_index)
    
    User_data = df.iloc[start_index:, [0, 5]].values.tolist()

    current_user_name = None

    for u in User_data[:303]:
        user_id += 1
        start_time = time.time()
        user_origin = file_number
        user_nickname = u[0]
        user_URL = u[1]

        current_user_name = user_nickname

        driver.get(user_URL)
        driver.implicitly_wait(5)
        if user_URL in user_URL_list:
            user_id -= 1
            print("user 중복", user_nickname, user_URL)
            continue
        user_URL_list.append(user_URL)
        try:
            photo_video_review_checkbox = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'onlyHasMedia')))
            if photo_video_review_checkbox.is_selected():
                driver.execute_script("arguments[0].click();", photo_video_review_checkbox)
            time.sleep(2)
        except Exception as e:
            print(f"체크박스가 존재하지 않거나 클릭 불가: {e}")
        first_post = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.FypLVg.qXYuuA')))

        driver.execute_script("arguments[0].click();", first_post)
        time.sleep(5) 


        for i in range(0,40):
            for c in range(0,30):
                driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
            time.sleep(1)


        try:
            nickname = driver.find_element(By.CSS_SELECTOR, 'h1._6MtIQ_').text.strip()

            if nickname != user_nickname:
                print(nickname, user_nickname, "같지 않음")

            stores = driver.find_elements(By.CSS_SELECTOR, 'div.eEfPp0')
            y = 0
            for s in stores:
                i_store_review = s.find_elements(By.CSS_SELECTOR, 'div.pui__vn15t2')
            
                if i_store_review and not i_store_review[0].text.strip():
                    store_review = "없음"
                else:
                    store_review = s.find_element(By.CSS_SELECTOR, 'div.pui__vn15t2').text.strip()

            
                if store_review != "없음":
                    store_name = s.find_element(By.CSS_SELECTOR, 'span.pui__pv1E2a').text.strip()

                    i_store_category = s.find_elements(By.CSS_SELECTOR, 'div.pui__Vb-OW1')

                    if i_store_category and not i_store_category[0].text.strip():
                        store_category = "없음"
                        continue
                    elif any("폐업" in element.text for element in i_store_category):
                        store_category = "폐업했거나 정보 제공이 중지된 장소"
                        continue
                    else:
                        store_category_loc = s.find_elements(By.CSS_SELECTOR, 'span.pui__WUm6H8')
                        store_category = store_category_loc[0].text.strip()
                        store_loc = store_category_loc[1].text.strip()
                        
                        if "서울 서대문구" in store_loc:
                            if store_category not in etc:
                                i_store_date = s.find_element(By.CSS_SELECTOR, 'span.pui__gfuUIT').text.strip()
                                line = i_store_date.split('\n')
                                store_date = '2025' + line[-1] if line[-1][0] == '년' else '20' + line[-1]

                                revisit_elements = s.find_elements(By.CSS_SELECTOR, 'span.pui__gfuUIT')
                                revisit = revisit_elements[1].text.strip() if len(revisit_elements) > 1 else ''

                                store_hour_element = s.find_elements(By.CSS_SELECTOR, 'span.pui__V8F9nN.pui__2ZezJb > em')
                                store_hour = store_hour_element[0].text.strip() if store_hour_element else ''
                                
                                # revisit "번째 방문" 문자 제거
                                if revisit:
                                    revisit = revisit.replace(',', '')
                                    revisit = int(revisit[:-5])
                                    
                                if store_hour:
                                    store_hour = store_hour[:-4]
                            else:
                                continue
                        else:
                            continue
                                    
                        list_sheet.append([user_id, nickname, file_number, store_name, store_loc, store_category, store_review, store_date, revisit, store_hour, user_URL])
                        y = y+1
            end_time = time.time()
            elapsed_time = int(end_time - start_time)
            xlsx.save(file_name)
            print(f"number = {y}, (time = {elapsed_time})")

        except Exception as e:
            print('Error:', e)
    xlsx.save(file_name)
    print("list_sheet", list_sheet)
    print(f"Data{file_number} end.")

xlsx.save(file_name)
driver.quit()

print("END!")